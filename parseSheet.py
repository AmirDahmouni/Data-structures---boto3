import openpyxl
from collections import defaultdict

def parseExcel(fileName):

  # Load the workbook and select the active sheet
  inv_file = openpyxl.load_workbook(fileName)

  product_list = inv_file["Sheet1"]

  # Use defaultdict to avoid key existence checks
  products_per_supplier = defaultdict(int)
  total_value_per_supplier = defaultdict(float)
  products_under_10_inv = {}

  # Iterate over the product rows, starting from the second row
  for product_row in range(2, product_list.max_row + 1):
      supplier_name = product_list.cell(product_row, 4).value
      inventory = product_list.cell(product_row, 2).value
      price = product_list.cell(product_row, 3).value
      product_num = product_list.cell(product_row, 1).value
      inventory_price = product_list.cell(product_row, 5)

      # Increment product count per supplier
      products_per_supplier[supplier_name] += 1

      # Add to the total value of inventory per supplier
      total_value_per_supplier[supplier_name] += inventory * price

      # Track products with inventory less than 10
      if inventory < 10:
          products_under_10_inv[int(product_num)] = int(inventory)

      # Update the total inventory price in the worksheet
      inventory_price.value = inventory * price

  # Print the results
  print(dict(products_per_supplier))
  print(dict(total_value_per_supplier))
  print(products_under_10_inv)

  # Save the workbook with the updated values
  inv_file.save("inventory_with_total_value.xlsx")


